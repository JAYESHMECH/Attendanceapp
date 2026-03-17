from flask import Flask, request, render_template
import openpyxl
from datetime import datetime

app = Flask(__name__)

file_name = "attendance.xlsx"

students = {
    "1": "Arpit Sarangdhar Dhok",
    "2": "Ayush Devchand Ganvir",
    "3": "Kaustubh Ravikant Ambulkar",
    "4": "Khushi Shyam Chorpagar",
    "5": "Khushi Suresh Warhade",
    "6": "Madhur Vivek Mamarde",
    "7": "Prerna Arun Thakare",
    "8": "Sameer Anis Beg",
    "9": "Shantatu Manish Athawale",
    "10": "Viashnavi Nandkumar Kokate",
    "11": "Yash Gajanan Khade",
    "12": "Yash Nilesh Borode",
    "13": "Yash Rajkumarji Futane",
    "14": "Nikita Santosh Thorat",
    "15": "Dhruv Rameshwar Pojage",
    "16": "Jayesh Satish Alaspure",
    "17": "Prajwal Nivrutti Metekar",
    "18": "Prathamesh Satish Murade",
    "19": "Priyanka Kailash Dhole",
    "20": "Akansha Ravindra Jondhale",
    "21": "Ashutosh Rajendra Araj",
    "22": "Avinash Sanjay Panzade",
    "23": "Ayush Laxmanrao Kadu",
    "24": "Dikshant Shivdas Raut",
    "25": "Payal Santosh Kakne",
    "26": "Ritika Sudish Dulgaj",
    "27": "Shubham Rajan Kawale",
    "28": "Arush Jivan Sadar",
    "29": "Harsh Pramod Khandare",
    "30": "Nidhi Vijay Lande",
    "31": "Om Nitin Men",
    "32": "Pranav Ravindra Ingale",
    "33": "Tejas Baban Nage",
    "34": "Yash Rajendra Chakravarti"
}

def create_file():
    try:
        openpyxl.load_workbook(file_name)
    except:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "Roll No", "Homework", "Attendance", "Date"])
        wb.save(file_name)

create_file()

@app.route('/')
def home():
    return render_template("index.html",students=students)

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    roll = request.form['roll']
    homework = request.form['homework']
    name = students.get(roll)
    if homework == "Yes":
        attendance = "Present"
    else:
        attendance = "Absent"

    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    sheet.append([name, roll, homework, attendance, datetime.now().strftime("%Y-%m-%d")])
    wb.save(file_name)

    return f"Attendance Marked: {attendance}"

if __name__=="__main__":
    app.run(debug=True)